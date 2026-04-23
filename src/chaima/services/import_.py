import csv as csv_module
import re
import uuid as uuid_pkg
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Literal

from sqlalchemy.orm import selectinload
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from chaima.models.import_log import ImportLog

_QU_RE = re.compile(r"^\s*(-?\d+(?:[.,]\d+)?)\s*([a-zA-ZµμÅ%°]+)?\s*$")


def split_quantity_unit(s: str) -> tuple[float | None, str | None]:
    if not s:
        return (None, None)
    m = _QU_RE.match(s)
    if m is None:
        return (None, None)
    qty_str = m.group(1).replace(",", ".")
    try:
        qty = float(qty_str)
    except ValueError:
        return (None, None)
    unit = m.group(2)
    return (qty, unit)


_HEADER_PATTERNS: list[tuple[str, str]] = [
    ("cas", "cas"),
    ("mit einheit", "quantity_unit_combined"),
    ("with unit", "quantity_unit_combined"),
    ("menge", "quantity"),
    ("quantity", "quantity"),
    ("amount", "quantity"),
    ("einheit", "unit"),
    ("unit", "unit"),
    ("standort", "location_text"),
    ("location", "location_text"),
    ("lagerort", "location_text"),
    ("lieferant", "supplier_text"),
    ("hersteller", "supplier_text"),
    ("supplier", "supplier_text"),
    ("vendor", "supplier_text"),
    ("reinheit", "purity"),
    ("purity", "purity"),
    ("bestellt von", "ordered_by"),
    ("ordered by", "ordered_by"),
    ("besteller", "ordered_by"),
    ("erstellt von", "created_by_name"),
    ("created by", "created_by_name"),
    ("label", "identifier"),
    ("behälter", "identifier"),
    ("identifier", "identifier"),
    ("bemerkung", "comment"),
    ("kommentar", "comment"),
    ("comment", "comment"),
    ("notes", "comment"),
    ("kaufdatum", "purchased_at"),
    ("gekauft", "purchased_at"),
    ("purchased", "purchased_at"),
    ("bought", "purchased_at"),
    ("name", "name"),
]


def detect_header_mapping(columns: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for col in columns:
        lower = col.strip().lower()
        chosen: str | None = None
        for pattern, target in _HEADER_PATTERNS:
            if pattern in lower:
                chosen = target
                break
        result[col] = chosen or "ignore"
    return result


@dataclass
class Grid:
    columns: list[str]
    rows: list[list[str]]
    row_count: int
    sheets: list[str] | None


def parse_upload(
    data: bytes, fmt: Literal["xlsx", "csv"], sheet_name: str | None = None
) -> Grid:
    if fmt == "xlsx":
        return _parse_xlsx(data, sheet_name)
    if fmt == "csv":
        return _parse_csv(data)
    raise ValueError(f"Unsupported format: {fmt}")


def _parse_xlsx(data: bytes, sheet_name: str | None) -> Grid:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {sheet_names}")
        ws = wb[sheet_name]

    all_rows = [
        [("" if cell is None else str(cell)) for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]
    if not all_rows:
        return Grid(columns=[], rows=[], row_count=0, sheets=sheet_names)

    columns = [str(c).strip() for c in all_rows[0]]
    rows = all_rows[1:]
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()
    return Grid(columns=columns, rows=rows, row_count=len(rows), sheets=sheet_names)


def _parse_csv(data: bytes) -> Grid:
    text = data.decode("utf-8-sig")
    reader = csv_module.reader(StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return Grid(columns=[], rows=[], row_count=0, sheets=None)
    columns = [c.strip() for c in all_rows[0]]
    rows = all_rows[1:]
    while rows and all(cell == "" for cell in rows[-1]):
        rows.pop()
    return Grid(columns=columns, rows=rows, row_count=len(rows), sheets=None)


@dataclass
class ParsedRow:
    index: int
    name: str | None
    cas: str | None
    location_text: str | None
    supplier_text: str | None
    quantity: float | None
    unit: str | None
    purity: str | None
    purchased_at: str | None
    ordered_by: str | None
    identifier: str | None
    created_by_name: str | None
    comment: str | None
    errors: list[str]
    warnings: list[str]


class MappingValidationError(ValueError):
    pass


_REQUIRED_TARGETS = {"name"}


def apply_column_mapping(
    grid: Grid,
    mapping: dict[str, str],
    qu_combined_column: str | None,
) -> list[ParsedRow]:
    targets_in_use = set(mapping.values())
    missing = _REQUIRED_TARGETS - targets_in_use
    if missing:
        raise MappingValidationError(f"Missing required columns: {sorted(missing)}")

    col_index = {col: i for i, col in enumerate(grid.columns)}
    parsed: list[ParsedRow] = []
    for i, row in enumerate(grid.rows):
        errors: list[str] = []
        warnings: list[str] = []
        values: dict[str, str | None] = {t: None for t in [
            "name", "cas", "location_text", "supplier_text", "quantity", "unit",
            "purity", "purchased_at", "ordered_by", "identifier",
            "created_by_name", "comment",
        ]}
        qty: float | None = None
        unit: str | None = None
        for source_col, target in mapping.items():
            if target == "ignore":
                continue
            cell = row[col_index[source_col]] if col_index[source_col] < len(row) else ""
            cell = cell.strip() if cell else ""
            if target in ("quantity_unit_combined", "quantity"):
                if cell:
                    q, u = _parse_qty(cell)
                    qty = q if q is not None else qty
                    if u is not None:
                        unit = u
                    if q is None:
                        warnings.append(f"Could not parse quantity from '{cell}'")
            elif target == "unit":
                unit = cell or None
            else:
                values[target] = cell or None

        parsed.append(ParsedRow(
            index=i,
            name=values["name"],
            cas=values["cas"],
            location_text=values["location_text"],
            supplier_text=values["supplier_text"],
            quantity=qty,
            unit=unit,
            purity=values["purity"],
            purchased_at=values["purchased_at"],
            ordered_by=values["ordered_by"],
            identifier=values["identifier"],
            created_by_name=values["created_by_name"],
            comment=values["comment"],
            errors=errors,
            warnings=warnings,
        ))
    return parsed


_NUM_RE = re.compile(r"(-?\d+(?:[.,]\d+)?)")


def _parse_qty(cell: str) -> tuple[float | None, str | None]:
    try:
        return (float(cell.replace(",", ".")), None)
    except ValueError:
        pass
    q, u = split_quantity_unit(cell)
    if q is not None:
        return (q, u)
    m = _NUM_RE.search(cell)
    if m is None:
        return (None, None)
    num_str = m.group(1).replace(",", ".")
    rest = re.sub(r"[\d.,]", "", cell).strip()
    try:
        return (float(num_str), rest or None)
    except ValueError:
        return (None, None)


@dataclass
class ChemicalGroup:
    canonical_name: str
    canonical_cas: str | None
    row_indices: list[int]


def _normalize_name(s: str | None) -> str:
    return (s or "").strip().lower()


def group_chemicals_by_identity(rows: list[ParsedRow]) -> list[ChemicalGroup]:
    buckets: dict[str, list[ParsedRow]] = {}
    for r in rows:
        if r.cas:
            key = f"cas:{r.cas}"
        else:
            key = f"name:{_normalize_name(r.name)}"
        buckets.setdefault(key, []).append(r)

    groups: list[ChemicalGroup] = []
    for key, rs in buckets.items():
        canonical = next((r.name for r in rs if r.name), rs[0].name or "")
        canonical_cas = next((r.cas for r in rs if r.cas), None)
        groups.append(ChemicalGroup(
            canonical_name=canonical,
            canonical_cas=canonical_cas,
            row_indices=[r.index for r in rs],
        ))
    return groups


@dataclass
class LocationMapping:
    source_text: str
    location_id: uuid_pkg.UUID | None
    new_location: dict | None


@dataclass
class ChemicalGroupPayload:
    canonical_name: str
    canonical_cas: str | None
    row_indices: list[int]


@dataclass
class CommitPayload:
    column_mapping: dict[str, str]
    quantity_unit_combined_column: str | None
    columns: list[str]
    rows: list[list[str]]
    location_mapping: list[LocationMapping]
    chemical_groups: list[ChemicalGroupPayload]


@dataclass
class ImportSummary:
    created_chemicals: int
    created_containers: int
    created_locations: int
    skipped_rows: list[dict]
    warnings: list[dict]


class ImportValidationError(ValueError):
    def __init__(self, errors: list[dict]):
        self.errors = errors
        super().__init__(f"Import validation failed: {len(errors)} error(s)")


async def commit_import(
    session,
    *,
    group_id: uuid_pkg.UUID,
    viewer_id: uuid_pkg.UUID,
    payload: CommitPayload,
) -> ImportSummary:
    from chaima.models.chemical import Chemical
    from chaima.models.storage import StorageKind, StorageLocation, StorageLocationGroup
    from chaima.services import containers as container_service

    parsed = apply_column_mapping(
        Grid(columns=payload.columns, rows=payload.rows, row_count=len(payload.rows), sheets=None),
        payload.column_mapping,
        payload.quantity_unit_combined_column,
    )
    blank_indices: set[int] = set()
    real_errors: list[dict] = []
    row_warnings: list[dict] = []
    for r in parsed:
        if not r.name:
            blank_indices.add(r.index)
            continue
        if r.errors:
            real_errors.append({"index": r.index, "reason": "; ".join(r.errors)})
        if r.warnings:
            row_warnings.append({
                "chemical": r.name,
                "row": r.index + 1,
                "details": "; ".join(r.warnings),
            })
    if real_errors:
        raise ImportValidationError(real_errors)
    parsed = [r for r in parsed if r.index not in blank_indices]

    location_text_to_id: dict[str, uuid_pkg.UUID] = {}
    created_locations = 0
    for lm in payload.location_mapping:
        if lm.location_id is not None:
            location_text_to_id[lm.source_text] = lm.location_id
        elif lm.new_location is not None:
            new_loc = StorageLocation(
                name=lm.new_location["name"],
                kind=StorageKind.CABINET,
                parent_id=lm.new_location.get("parent_id"),
            )
            session.add(new_loc)
            await session.flush()
            session.add(StorageLocationGroup(location_id=new_loc.id, group_id=group_id))
            await session.flush()
            location_text_to_id[lm.source_text] = new_loc.id
            created_locations += 1
        else:
            raise ImportValidationError(
                [{"index": -1, "reason": f"Location '{lm.source_text}' has no mapping"}]
            )

    from sqlmodel import select

    existing_chems = (await session.exec(
        select(Chemical).where(Chemical.group_id == group_id)
    )).all()
    chem_by_name: dict[str, uuid_pkg.UUID] = {
        c.name.strip().lower(): c.id for c in existing_chems
    }

    row_to_chemical: dict[int, uuid_pkg.UUID] = {}
    created_chemicals = 0
    for cg in payload.chemical_groups:
        if not (cg.canonical_name or "").strip():
            continue
        key = cg.canonical_name.strip().lower()
        existing_id = chem_by_name.get(key)
        if existing_id:
            chem_id = existing_id
        else:
            chem = Chemical(
                name=cg.canonical_name,
                cas=cg.canonical_cas,
                group_id=group_id,
                created_by=viewer_id,
            )
            session.add(chem)
            await session.flush()
            chem_id = chem.id
            chem_by_name[key] = chem_id
            created_chemicals += 1
        for idx in cg.row_indices:
            row_to_chemical[idx] = chem_id

    from chaima.models.supplier import Supplier

    supplier_cache: dict[str, uuid_pkg.UUID] = {}
    existing_suppliers = (await session.exec(
        select(Supplier).where(Supplier.group_id == group_id)
    )).all()
    for s in existing_suppliers:
        supplier_cache[s.name.strip().lower()] = s.id

    default_location_id: uuid_pkg.UUID | None = None
    created_suppliers = 0

    created_containers = 0
    for row in parsed:
        chem_id = row_to_chemical.get(row.index)
        if chem_id is None:
            continue
        loc_id = location_text_to_id.get(row.location_text or "") if row.location_text else None
        if loc_id is None:
            if default_location_id is None:
                default_loc = StorageLocation(
                    name="Imported",
                    kind=StorageKind.CABINET,
                )
                session.add(default_loc)
                await session.flush()
                session.add(StorageLocationGroup(location_id=default_loc.id, group_id=group_id))
                await session.flush()
                default_location_id = default_loc.id
                created_locations += 1
            loc_id = default_location_id

        supplier_id: uuid_pkg.UUID | None = None
        if row.supplier_text:
            key = row.supplier_text.strip().lower()
            if key in supplier_cache:
                supplier_id = supplier_cache[key]
            else:
                new_sup = Supplier(name=row.supplier_text.strip(), group_id=group_id)
                session.add(new_sup)
                await session.flush()
                supplier_cache[key] = new_sup.id
                supplier_id = new_sup.id
                created_suppliers += 1

        identifier = row.identifier or _next_identifier(row.name or "X")
        try:
            container = await container_service.create_container(
                session,
                chemical_id=chem_id,
                location_id=loc_id,
                identifier=identifier,
                amount=row.quantity if row.quantity is not None else 0.0,
                unit=row.unit or "",
                created_by=viewer_id,
                supplier_id=supplier_id,
            )
        except container_service.DuplicateIdentifier:
            identifier = f"{identifier}-{row.index}"
            container = await container_service.create_container(
                session,
                chemical_id=chem_id,
                location_id=loc_id,
                identifier=identifier,
                amount=row.quantity if row.quantity is not None else 0.0,
                unit=row.unit or "",
                created_by=viewer_id,
                supplier_id=supplier_id,
            )
        if row.ordered_by:
            container.ordered_by_name = row.ordered_by
        if row.purity:
            container.purity = row.purity
        created_containers += 1

    await session.flush()
    return ImportSummary(
        created_chemicals=created_chemicals,
        created_containers=created_containers,
        created_locations=created_locations,
        skipped_rows=[],
        warnings=row_warnings,
    )


_identifier_counters: dict[str, int] = {}


def _next_identifier(chem_name: str) -> str:
    key = chem_name[:1].upper() if chem_name else "X"
    _identifier_counters[key] = _identifier_counters.get(key, 0) + 1
    return f"{key}-IMP-{_identifier_counters[key]:04d}"


async def find_previous_import(
    session: AsyncSession,
    *,
    group_id: uuid_pkg.UUID,
    file_name: str,
) -> ImportLog | None:
    if not file_name:
        return None
    result = await session.exec(
        select(ImportLog)
        .where(ImportLog.group_id == group_id, ImportLog.file_name == file_name)
        .options(selectinload(ImportLog.user))  # type: ignore[arg-type]
        .order_by(ImportLog.created_at.desc())  # type: ignore[union-attr]
        .limit(1)
    )
    return result.first()


async def log_import(
    session: AsyncSession,
    *,
    group_id: uuid_pkg.UUID,
    file_name: str,
    imported_by: uuid_pkg.UUID,
    row_count: int,
) -> ImportLog:
    entry = ImportLog(
        group_id=group_id,
        file_name=file_name,
        imported_by=imported_by,
        row_count=row_count,
    )
    session.add(entry)
    await session.flush()
    return entry
