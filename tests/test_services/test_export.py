from io import StringIO, BytesIO
from uuid import uuid4
import csv

from openpyxl import load_workbook

from chaima.models.chemical import Chemical
from chaima.models.container import Container
from chaima.models.storage import StorageLocation
from chaima.models.supplier import Supplier
from chaima.services import export as export_service


async def test_export_csv_one_row_per_container(session, group, user):
    loc = StorageLocation(name="Shelf A", kind="shelf")
    session.add(loc)
    sup = Supplier(name="Sigma", group_id=group.id)
    session.add(sup)
    chem = Chemical(name="Ethanol", cas="64-17-5", group_id=group.id, created_by=user.id)
    session.add(chem)
    await session.flush()
    session.add(Container(
        chemical_id=chem.id, location_id=loc.id, supplier_id=sup.id,
        created_by=user.id, identifier="E-001", amount=1.0, unit="L",
        ordered_by_name="M. Schmidt",
    ))
    session.add(Container(
        chemical_id=chem.id, location_id=loc.id,
        created_by=user.id, identifier="E-002", amount=0.5, unit="L",
    ))
    await session.commit()

    data = await export_service.export_chemicals(session, group.id, viewer_id=user.id, filters={}, fmt="csv")
    reader = csv.reader(StringIO(data.decode("utf-8")))
    rows = list(reader)
    header = rows[0]
    body = rows[1:]

    assert header[:7] == ["name", "cas", "smiles", "location", "identifier", "quantity", "unit"]
    assert len(body) == 2
    e001 = next(r for r in body if r[4] == "E-001")
    assert e001[0] == "Ethanol"
    assert e001[1] == "64-17-5"
    assert e001[3] == "Shelf A"
    assert e001[5] == "1.0"
    assert e001[6] == "L"
    assert "M. Schmidt" in e001  # ordered_by column
    assert "Sigma" in e001        # supplier column


async def test_export_includes_chemical_without_containers(session, group, user):
    chem = Chemical(name="Isolated", group_id=group.id, created_by=user.id)
    session.add(chem)
    await session.commit()

    data = await export_service.export_chemicals(session, group.id, viewer_id=user.id, filters={}, fmt="csv")
    reader = csv.reader(StringIO(data.decode("utf-8")))
    rows = list(reader)
    body = rows[1:]
    assert len(body) == 1
    assert body[0][0] == "Isolated"
    assert body[0][4] == ""  # identifier empty
    assert body[0][5] == ""  # quantity empty


async def test_export_xlsx_round_trip(session, group, user):
    loc = StorageLocation(name="Shelf A", kind="shelf")
    session.add(loc)
    chem = Chemical(name="Ethanol", group_id=group.id, created_by=user.id)
    session.add(chem)
    await session.flush()
    session.add(Container(
        chemical_id=chem.id, location_id=loc.id, created_by=user.id,
        identifier="E-001", amount=1.0, unit="L",
    ))
    await session.commit()

    data = await export_service.export_chemicals(session, group.id, viewer_id=user.id, filters={}, fmt="xlsx")
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header[:5] == ["name", "cas", "smiles", "location", "identifier"]
    assert ws.cell(row=2, column=1).value == "Ethanol"


async def test_export_respects_location_filter(session, group, user):
    loc_a = StorageLocation(name="Shelf A", kind="shelf")
    loc_b = StorageLocation(name="Shelf B", kind="shelf")
    session.add(loc_a)
    session.add(loc_b)
    chem_a = Chemical(name="OnA", group_id=group.id, created_by=user.id)
    chem_b = Chemical(name="OnB", group_id=group.id, created_by=user.id)
    session.add(chem_a)
    session.add(chem_b)
    await session.flush()
    session.add(Container(chemical_id=chem_a.id, location_id=loc_a.id,
                          created_by=user.id, identifier="A-1", amount=1, unit="L"))
    session.add(Container(chemical_id=chem_b.id, location_id=loc_b.id,
                          created_by=user.id, identifier="B-1", amount=1, unit="L"))
    await session.commit()

    data = await export_service.export_chemicals(
        session, group.id, viewer_id=user.id, filters={"location_id": loc_a.id}, fmt="csv"
    )
    reader = csv.reader(StringIO(data.decode("utf-8")))
    body = list(reader)[1:]
    names = {r[0] for r in body}
    assert names == {"OnA"}


async def test_export_too_large_raises(session, group, user, monkeypatch):
    monkeypatch.setattr(export_service, "EXPORT_ROW_CAP", 1)
    chem = Chemical(name="A", group_id=group.id, created_by=user.id)
    chem2 = Chemical(name="B", group_id=group.id, created_by=user.id)
    session.add(chem)
    session.add(chem2)
    await session.commit()
    import pytest
    with pytest.raises(export_service.ExportTooLargeError):
        await export_service.export_chemicals(session, group.id, viewer_id=user.id, filters={}, fmt="csv")


def test_export_columns_include_sds_url():
    from chaima.services.export import EXPORT_COLUMNS
    assert "sds_url" in EXPORT_COLUMNS
    assert EXPORT_COLUMNS.index("sds_url") == EXPORT_COLUMNS.index("comment") + 1


def test_row_for_container_includes_sds_url():
    from chaima.services.export import EXPORT_COLUMNS, _row_for_container

    chem = Chemical(
        name="ExpChem", group_id=uuid4(), created_by=uuid4(),
        sds_url="https://example.com/sds.pdf",
    )
    chem.ghs_links = []
    chem.hazard_tag_links = []
    row = _row_for_container(chem, None)
    assert len(row) == len(EXPORT_COLUMNS)
    assert row[EXPORT_COLUMNS.index("sds_url")] == "https://example.com/sds.pdf"


def test_row_for_container_without_sds_url_is_empty_string():
    from chaima.services.export import EXPORT_COLUMNS, _row_for_container

    chem = Chemical(name="NoUrlChem", group_id=uuid4(), created_by=uuid4())
    chem.ghs_links = []
    chem.hazard_tag_links = []
    row = _row_for_container(chem, None)
    assert len(row) == len(EXPORT_COLUMNS)
    assert row[EXPORT_COLUMNS.index("sds_url")] == ""
