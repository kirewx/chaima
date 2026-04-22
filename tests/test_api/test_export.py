from io import BytesIO, StringIO
import csv
from openpyxl import load_workbook

from chaima.models.chemical import Chemical
from chaima.models.container import Container
from chaima.models.storage import StorageLocation


async def test_export_csv_happy_path(client, session, group, user, membership):
    loc = StorageLocation(name="Shelf A", kind="shelf")
    session.add(loc)
    chem = Chemical(name="Ethanol", group_id=group.id, created_by=user.id)
    session.add(chem)
    await session.flush()
    session.add(Container(
        chemical_id=chem.id, location_id=loc.id, created_by=user.id,
        identifier="E-001", amount=1.0, unit="L",
    ))
    await session.commit()

    resp = await client.get(
        f"/api/v1/groups/{group.id}/chemicals/export?format=csv"
    )
    assert resp.status_code == 200
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith(".csv\"")
    rows = list(csv.reader(StringIO(resp.text)))
    assert rows[0][0] == "name"
    assert rows[1][0] == "Ethanol"


async def test_export_xlsx_happy_path(client, session, group, user, membership):
    chem = Chemical(name="Ethanol", group_id=group.id, created_by=user.id)
    session.add(chem)
    await session.commit()

    resp = await client.get(
        f"/api/v1/groups/{group.id}/chemicals/export?format=xlsx"
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "name"
    assert ws.cell(row=2, column=1).value == "Ethanol"


async def test_export_unknown_format_400(client, group, membership):
    resp = await client.get(
        f"/api/v1/groups/{group.id}/chemicals/export?format=pdf"
    )
    assert resp.status_code == 422  # pydantic Literal validator


async def test_export_respects_has_containers_filter(client, session, group, user, membership):
    loc = StorageLocation(name="Shelf A", kind="shelf")
    session.add(loc)
    chem_empty = Chemical(name="Empty", group_id=group.id, created_by=user.id)
    chem_full = Chemical(name="Full", group_id=group.id, created_by=user.id)
    session.add(chem_empty)
    session.add(chem_full)
    await session.flush()
    session.add(Container(
        chemical_id=chem_full.id, location_id=loc.id, created_by=user.id,
        identifier="F-1", amount=1, unit="L",
    ))
    await session.commit()

    resp = await client.get(
        f"/api/v1/groups/{group.id}/chemicals/export?format=csv&has_containers=true"
    )
    rows = list(csv.reader(StringIO(resp.text)))
    names = {r[0] for r in rows[1:]}
    assert names == {"Full"}


async def test_export_not_member_403(client, group):
    resp = await client.get(
        f"/api/v1/groups/{group.id}/chemicals/export?format=csv"
    )
    assert resp.status_code == 403
